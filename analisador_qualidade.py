import csv
import requests
import time
import os
import pandas as pd
from datetime import datetime
from typing import Dict, Any
from dotenv import load_dotenv

# Carrega as variáveis de ambiente
load_dotenv('config.env')

def verificar_qualidade_numero(cod_telefone: str, access_token: str) -> Dict[str, Any]:
    """
    Verifica a qualidade de um número de WhatsApp usando a API do Facebook Graph.
    
    Args:
        cod_telefone: O código do telefone para verificar
        access_token: Token de acesso do Facebook
        
    Returns:
        Dicionário com informações sobre a qualidade do número
    """
    url = f"https://graph.facebook.com/v23.0/{cod_telefone}"
    
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            
            # Verifica se há informações de qualidade no response
            if 'quality_rating' in data:
                quality = data['quality_rating']
                
                # Mapeia os status para português
                status_map = {
                    'GREEN': 'BOM',
                    'YELLOW': 'MÉDIO', 
                    'RED': 'RUIM'
                }
                
                status_pt = status_map.get(quality, 'DESCONHECIDO')
                
                return {
                    'status': 'SUCESSO',
                    'qualidade': quality,
                    'qualidade_pt': status_pt,
                    'dados': data
                }
            else:
                return {
                    'status': 'SEM_INFO_QUALIDADE',
                    'qualidade': 'N/A',
                    'qualidade_pt': 'SEM INFORMAÇÃO',
                    'dados': data
                }
                
        elif response.status_code == 404:
            return {
                'status': 'NAO_ENCONTRADO',
                'qualidade': 'N/A',
                'qualidade_pt': 'NÚMERO NÃO ENCONTRADO',
                'dados': None
            }
        elif response.status_code == 401:
            return {
                'status': 'ERRO_AUTENTICACAO',
                'qualidade': 'N/A',
                'qualidade_pt': 'ERRO DE AUTENTICAÇÃO - TOKEN INVÁLIDO',
                'dados': None
            }
        elif response.status_code == 400:
            return {
                'status': 'ERRO_REQUISICAO',
                'qualidade': 'N/A',
                'qualidade_pt': f'ERRO 400 - REQUISIÇÃO INVÁLIDA',
                'dados': None
            }
        else:
            return {
                'status': 'ERRO_API',
                'qualidade': 'N/A',
                'qualidade_pt': f'ERRO {response.status_code}',
                'dados': None
            }
            
    except requests.exceptions.RequestException as e:
        return {
            'status': 'ERRO_CONEXAO',
            'qualidade': 'N/A',
            'qualidade_pt': f'ERRO DE CONEXÃO: {str(e)}',
            'dados': None
        }

def gerar_relatorio_excel(resultados: list, nome_arquivo: str = None):
    """
    Gera um relatório em formato Excel (.xlsx) usando pandas.
    
    Args:
        resultados: Lista com os resultados da análise
        nome_arquivo: Nome do arquivo Excel (opcional)
    """
    if nome_arquivo is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_qualidade_{timestamp}.xlsx"
    
    # Cria DataFrame com os resultados
    df_data = []
    for resultado in resultados:
        df_data.append({
            'Nome': resultado.get('nome', ''),
            'Número': resultado.get('numero', ''),
            'Código Telefone': resultado.get('cod_telefone', ''),
            'Status': resultado.get('status', ''),
            'Qualidade': resultado.get('qualidade_pt', ''),
            'Data/Hora Verificação': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
    
    df = pd.DataFrame(df_data)
    
    # Cria um ExcelWriter para múltiplas abas
    with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
        # Aba 1: Dados completos
        df.to_excel(writer, sheet_name='Dados Completos', index=False)
        
        # Aplica formatação condicional na coluna Qualidade
        workbook = writer.book
        worksheet = writer.sheets['Dados Completos']
        
        # Define as cores para cada qualidade
        from openpyxl.styles import PatternFill
        
        verde = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Verde claro
        amarelo = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Amarelo claro
        vermelho = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Vermelho claro
        cinza = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Cinza para erros
        
        # Aplica cores na coluna Qualidade (coluna E, índice 4)
        for row in range(2, len(df) + 2):  # Começa da linha 2 (após cabeçalho)
            cell = worksheet.cell(row=row, column=5)  # Coluna E = Qualidade
            qualidade = cell.value
            
            if qualidade == 'BOM':
                cell.fill = verde
            elif qualidade == 'MÉDIO':
                cell.fill = amarelo
            elif qualidade == 'RUIM':
                cell.fill = vermelho
            else:
                cell.fill = cinza  # Para erros ou valores desconhecidos
        
        # Aba 2: Resumo por qualidade
        contadores = {'BOM': 0, 'MÉDIO': 0, 'RUIM': 0, 'ERROS': 0}
        for resultado in resultados:
            if resultado['status'] == 'SUCESSO':
                contadores[resultado['qualidade_pt']] += 1
            else:
                contadores['ERROS'] += 1
        
        resumo_data = {
            'Categoria': ['Números Bons', 'Números Médios', 'Números Ruins', 'Erros', 'Total'],
            'Quantidade': [
                contadores['BOM'],
                contadores['MÉDIO'],
                contadores['RUIM'],
                contadores['ERROS'],
                len(resultados)
            ],
            'Percentual': [
                f"{(contadores['BOM']/len(resultados)*100):.1f}%",
                f"{(contadores['MÉDIO']/len(resultados)*100):.1f}%",
                f"{(contadores['RUIM']/len(resultados)*100):.1f}%",
                f"{(contadores['ERROS']/len(resultados)*100):.1f}%",
                "100%"
            ]
        }
        
        df_resumo = pd.DataFrame(resumo_data)
        df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
        
        # Aba 3: Números por qualidade
        for qualidade in ['BOM', 'MÉDIO', 'RUIM']:
            numeros_qualidade = [r for r in resultados if r.get('qualidade_pt') == qualidade]
            if numeros_qualidade:
                df_qualidade = pd.DataFrame([{
                    'Nome': r.get('nome', ''),
                    'Número': r.get('numero', ''),
                    'Código': r.get('cod_telefone', ''),
                    'Qualidade': r.get('qualidade_pt', '')
                } for r in numeros_qualidade])
                
                df_qualidade.to_excel(writer, sheet_name=f'Números {qualidade}', index=False)
                
                # Aplica cor na aba específica
                worksheet_qualidade = writer.sheets[f'Números {qualidade}']
                cor_fill = None
                
                if qualidade == 'BOM':
                    cor_fill = verde
                elif qualidade == 'MÉDIO':
                    cor_fill = amarelo
                elif qualidade == 'RUIM':
                    cor_fill = vermelho
                
                if cor_fill:
                    # Aplica cor em toda a coluna Qualidade da aba específica
                    for row in range(2, len(df_qualidade) + 2):
                        cell = worksheet_qualidade.cell(row=row, column=4)  # Coluna D = Qualidade
                        cell.fill = cor_fill
        
        # Aba 4: Erros (se houver)
        erros = [r for r in resultados if r['status'] != 'SUCESSO']
        if erros:
            df_erros = pd.DataFrame([{
                'Nome': r.get('nome', ''),
                'Número': r.get('numero', ''),
                'Código': r.get('cod_telefone', ''),
                'Tipo de Erro': r.get('qualidade_pt', '')
            } for r in erros])
            
            df_erros.to_excel(writer, sheet_name='Erros', index=False)
            
            # Aplica cor cinza na aba de erros
            worksheet_erros = writer.sheets['Erros']
            for row in range(2, len(df_erros) + 2):
                cell = worksheet_erros.cell(row=row, column=4)  # Coluna D = Tipo de Erro
                cell.fill = cinza
        
        # Aba 5: Informações do relatório
        info_data = {
            'Informação': [
                'Data/Hora de Geração',
                'Total de Números Analisados',
                'Números com Sucesso',
                'Números com Erro',
                'Taxa de Sucesso'
            ],
            'Valor': [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                len(resultados),
                len([r for r in resultados if r['status'] == 'SUCESSO']),
                len([r for r in resultados if r['status'] != 'SUCESSO']),
                f"{(len([r for r in resultados if r['status'] == 'SUCESSO'])/len(resultados)*100):.1f}%"
            ]
        }
        
        df_info = pd.DataFrame(info_data)
        df_info.to_excel(writer, sheet_name='Informações', index=False)
    
    print(f"\n📊 Relatório Excel gerado: {nome_arquivo}")
    return nome_arquivo



def analisar_numeros_csv(arquivo_csv: str = 'Números.csv'):
    """
    Analisa todos os números do arquivo CSV e verifica sua qualidade.
    
    Args:
        arquivo_csv: Caminho para o arquivo CSV
    """
    print("🔍 ANALISADOR DE QUALIDADE DOS NÚMEROS DE WHATSAPP")
    print("=" * 60)
    
    # Verifica se o access token está configurado
    access_token = os.getenv('FACEBOOK_ACCESS_TOKEN')
    if not access_token or access_token == 'seu_access_token_aqui':
        print("❌ ERRO: Access token não configurado!")
        print("📝 Por favor, edite o arquivo 'config.env' e adicione seu access token:")
        print("   FACEBOOK_ACCESS_TOKEN=seu_token_real_aqui")
        return
    
    try:
        with open(arquivo_csv, 'r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            
            # Lista para armazenar todos os resultados
            resultados = []
            
            for i, row in enumerate(reader, 1):
                nome = row['Nome'].strip()
                numero = row['Número'].strip()
                cod_telefone = row['Cod Telefone'].strip()
                
                print(f"\n📱 Verificando {i}: {nome}")
                print(f"   Número: {numero}")
                print(f"   Código: {cod_telefone}")
                
                # Verifica a qualidade do número
                resultado = verificar_qualidade_numero(cod_telefone, access_token)
                
                # Adiciona informações do número ao resultado
                resultado['nome'] = nome
                resultado['numero'] = numero
                resultado['cod_telefone'] = cod_telefone
                
                # Exibe o resultado
                if resultado['status'] == 'SUCESSO':
                    emoji = {
                        'BOM': '🟢',
                        'MÉDIO': '🟡', 
                        'RUIM': '🔴'
                    }.get(resultado['qualidade_pt'], '❓')
                    
                    print(f"   {emoji} Status: {resultado['qualidade_pt']} ({resultado['qualidade']})")
                else:
                    print(f"   ❌ {resultado['qualidade_pt']}")
                
                resultados.append(resultado)
                
                # Pequena pausa para não sobrecarregar a API
                time.sleep(1)
            
            # Resumo final
            print("\n" + "=" * 60)
            print("📊 RESUMO FINAL")
            print("=" * 60)
            
            # Conta os status
            contadores = {'BOM': 0, 'MÉDIO': 0, 'RUIM': 0, 'ERROS': 0}
            
            for resultado in resultados:
                if resultado['status'] == 'SUCESSO':
                    contadores[resultado['qualidade_pt']] += 1
                else:
                    contadores['ERROS'] += 1
            
            print(f"🟢 Números BONS: {contadores['BOM']}")
            print(f"🟡 Números MÉDIOS: {contadores['MÉDIO']}")
            print(f"🔴 Números RUINS: {contadores['RUIM']}")
            print(f"❌ Erros: {contadores['ERROS']}")
            print(f"📱 Total analisado: {len(resultados)}")
            
            # Lista detalhada por qualidade
            print("\n📋 DETALHAMENTO POR QUALIDADE:")
            print("-" * 40)
            
            for qualidade in ['BOM', 'MÉDIO', 'RUIM']:
                numeros_qualidade = [r for r in resultados if r.get('qualidade_pt') == qualidade]
                if numeros_qualidade:
                    emoji = {'BOM': '🟢', 'MÉDIO': '🟡', 'RUIM': '🔴'}[qualidade]
                    print(f"\n{emoji} {qualidade}:")
                    for num in numeros_qualidade:
                        print(f"   • {num['nome']} ({num['numero']})")
            
            # Lista de erros
            erros = [r for r in resultados if r['status'] != 'SUCESSO']
            if erros:
                print(f"\n❌ ERROS ENCONTRADOS:")
                for erro in erros:
                    print(f"   • {erro['nome']}: {erro['qualidade_pt']}")
            
            # Gera relatório
            print("\n" + "=" * 60)
            print("📄 GERANDO RELATÓRIO...")
            print("=" * 60)
            
            # Gera relatório Excel
            arquivo_excel = gerar_relatorio_excel(resultados)
            
            print(f"\n✅ Relatório salvo com sucesso:")
            print(f"   📊 Excel: {arquivo_excel}")
                    
    except FileNotFoundError:
        print(f"❌ Erro: Arquivo '{arquivo_csv}' não encontrado!")
    except Exception as e:
        print(f"❌ Erro inesperado: {str(e)}")

if __name__ == "__main__":
    analisar_numeros_csv() 